from typing import List, Generator, Union, Dict, Any
import json
import csv
import chardet
from io import StringIO,BytesIO
import pandas as pd
import docx
import pdfplumber
import xml.etree.ElementTree as ET
import zipfile
import openpyxl
import email
from libs.case_processor import CaseProcessor
import re

import logging
logger = logging.getLogger(__name__)

class UniversalDocumentReader:
    def __init__(self, file_content: bytes, filename: str, file: Any):
        self.content = file_content
        self.filename = filename.lower()
        self.text_content = None
        self.case_processor = CaseProcessor()
        self.file = file
    
    def detect_encoding(self) -> str:
        result = chardet.detect(self.content)
        return result['encoding'] or 'utf-8'
    
    def read_text_file(self) -> Generator[str, None, None]:
        encoding = self.detect_encoding()
        try:
            text = self.content.decode(encoding)
            for line in text.splitlines():
                yield line.strip()
        except UnicodeDecodeError as e:
            yield f"Error decoding file: {str(e)}"
    
    def read_json_file(self) -> Generator[str, None, None]:
        try:
            encoding = self.detect_encoding()
            text = self.content.decode(encoding)
            data = json.loads(text)
            
            def process_json(obj, parent_key=''):
                if isinstance(obj, dict):
                    for key, value in obj.items():
                        new_key = f"{parent_key}.{key}" if parent_key else key
                        if isinstance(value, (dict, list)):
                            yield from process_json(value, new_key)
                        else:
                            yield f"{new_key}: {value}"
                elif isinstance(obj, list):
                    for i, item in enumerate(obj):
                        new_key = f"{parent_key}[{i}]"
                        if isinstance(item, (dict, list)):
                            yield from process_json(item, new_key)
                        else:
                            yield f"{new_key}: {item}"
                else:
                    yield f"{parent_key}: {obj}"
            
            yield from process_json(data)
        except json.JSONDecodeError as e:
            yield f"Error parsing JSON: {str(e)}"
    
    def read_csv_file(self) -> Generator[str, None, None]:

        try:
            encoding = self.detect_encoding()
            text = self.content.decode(encoding)
            csv_reader = csv.reader(StringIO(text))
            for row in csv_reader:
                yield ', '.join(row)
        except Exception as e:
            yield f"Error reading CSV: {str(e)}"
    
    def read_excel_file(self) -> Generator[str, None, None]:

        try:
            wb = openpyxl.load_workbook(filename=BytesIO(self.content), data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                yield f"Sheet: {sheet}"
                for row in ws.iter_rows(values_only=True):
                    yield ', '.join(str(cell) for cell in row if cell is not None)
        except Exception as e:
            yield f"Error reading Excel file: {str(e)}"

    def is_headline(self,line):
        words = line.strip().split()
        if not words:
            return False

  
        content_starters = ["A ", "The ", "In ", "And ", "But "]
        if any(line.startswith(starter) for starter in content_starters):
            return False

   
        if any(x in line for x in ['"', '@', '+1-', '(', ')']):
            return False

      
        small_words = {'a', 'an', 'and', 'as', 'at', 'but', 'by', 'for', 'in', 'of', 'on', 'or', 'the', 'to', 'with'}
        for word in words:
            if word.lower() not in small_words and not word[0].isupper():
                return False

        return True
    
    def read_pdf_file(self) -> Dict[str, Any]:
        try:
            cases = {}
            
            with pdfplumber.open(self.file) as pdf:
                for i,page in enumerate(pdf.pages):
                    text = page.extract_text()
                    lines = [line.strip() for line in text.splitlines() if line.strip()]
                    
                    if not text.strip():
                        continue

                    head=''
                    content=[]
                    for j,line in enumerate(lines):
                        if self.is_headline(line):
                            if j==0:
                                head+=line
                            else:
                                head+=" "+line
                        else:
                            content.append(line)

                    cases[i+1]= {
                        "headline":head,
                        "content":content
                    }

                    # cases = self.case_processor.process_document(lines)
                    
                
            if not cases:
                raise ValueError("No valid content found in PDF")
                
            return {
                'document_type': 'PDF',
                'total_pages': len(cases),
                'cases': cases
            }
            
        except Exception as e:
            logger.error(f"Error reading PDF: {str(e)}")
            raise
    
    def read_docx_file(self) -> Generator[str, None, None]:

        try:
            doc = docx.Document(BytesIO(self.content))
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    yield paragraph.text.strip()
        except Exception as e:
            yield f"Error reading DOCX: {str(e)}"
    
    def read_xml_file(self) -> Generator[str, None, None]:
        try:
            root = ET.fromstring(self.content)
            def process_element(element, level=0):
                indent = "  " * level
                if element.text and element.text.strip():
                    yield f"{indent}{element.tag}: {element.text.strip()}"
                for child in element:
                    yield from process_element(child, level + 1)
            yield from process_element(root)
        except ET.ParseError as e:
            yield f"Error parsing XML: {str(e)}"
    
    def read_email_file(self) -> Generator[str, None, None]:

        try:
            msg = email.message_from_bytes(self.content)
            yield f"Subject: {msg['subject']}"
            yield f"From: {msg['from']}"
            yield f"To: {msg['to']}"
            yield f"Date: {msg['date']}"
            
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain":
                        text = part.get_payload(decode=True).decode()
                        for line in text.splitlines():
                            if line.strip():
                                yield line.strip()
            else:
                text = msg.get_payload(decode=True).decode()
                for line in text.splitlines():
                    if line.strip():
                        yield line.strip()
        except Exception as e:
            yield f"Error reading email: {str(e)}"
    
    def read_lines(self) -> Generator[str, None, None]:
        file_readers = {
            '.txt': self.read_text_file,
            '.json': self.read_json_file,
            '.csv': self.read_csv_file,
            '.xlsx': self.read_excel_file,
            '.xls': self.read_excel_file,
            '.pdf': self.read_pdf_file,
            '.docx': self.read_docx_file,
            '.xml': self.read_xml_file,
            '.eml': self.read_email_file
        }
        
        ext = '.' + self.filename.split('.')[-1] if '.' in self.filename else ''
        
        reader = file_readers.get(ext, self.read_text_file)
        try:
            yield from reader()
        except Exception as e:
            yield f"Error processing file: {str(e)}"

    def process_document(self) -> Dict[str, Any]:
        try:
            if self.filename.endswith('.pdf'):
                return self.read_pdf_file()
                
            lines = list(self.read_lines())
            if not lines:
                raise ValueError(f"No content found in file: {self.filename}")
                
            cases = self.case_processor.process_document(lines)
            
            for case in cases:
                case['page_number'] = 1
                
            return {
                'document_type': self.filename.split('.')[-1].upper(),
                'total_pages': 1,
                'cases': cases
            }
            
        except Exception as e:
            logger.error(f"Error processing document: {str(e)}")
            raise
